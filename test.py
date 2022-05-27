# from openpyxl import load_workbook, Workbook
# from openpyxl.drawing.image import Image
# import numpy as np
#
#
# def inset_image(image_path, excel_path, image_location, i):
#     """
#     :param image_path: the image path
#     :param excel_path: the excel save path/read path
#     :param image_location: the location as image insert
#     """
#     image = Image(image_path)
#     image = scale_image(image, [4, 4])
#     # try:
#     wb = load_workbook(excel_path)
#     print(222)
#     ws = wb[i]
#     print(ws)
#     ws.add_image(image, image_location)
#     wb.save(r'C:\Users\txd\Desktop\out.xlsx')
#     # except:
#     #     wb = Workbook()
#     #     ws = wb.active
#     #     for i in range(5):
#     #         ws.title = 'sheet%s' % i
#     #         # pr
#     #         ws.add_image(image, image_location)
#     #         wb.save(r'C:\Users\txd\Desktop\out.xlsx')
#
#
def scale_image(image, size):
    """
    :param image: the image transform with openpyxl Image package
    :param size: the width and high scale size number
    :return: scale image
    """
    image_width = 720
    image_high = 1280
    img_scale_width = 0
    img_scale_high = 1

    image_size = (image_width / size[img_scale_width], image_high / size[img_scale_high])
    image.width, image.height = image_size
    return image

#
# if __name__ == "__main__":
#     wb = load_workbook(r'C:\Users\txd\Desktop\out (1).xlsx')
#     for i in wb:
#         inset_image(r'C:\Users\txd\Desktop\panxiuqiang.top.png', r'C:\Users\txd\Desktop\out (1).xlsx', 'A1', i)


# from openpyxl.rawing.image import Image\
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
wb = load_workbook(r'C:\Users\txd\Desktop\out (1).xlsx')  # 加载这个工作簿
for i in wb:
    print(i)
    print(i.title)
    sheet = wb[i.title]  # 选择你要操作的sheet表格
    img = Image(r'C:\Users\txd\Desktop\panxiuqiang.top.png')  # 选择你的图片
    image = scale_image(img, [4, 4])
    sheet.add_image(image, 'A1')
    wb.save(r'C:\Users\txd\Desktop\out (1).xlsx')  # 不要忘记保存
